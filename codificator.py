# Code to itarate over all the accounts extract the code and the account name
# python codificator.py
import openpyxl
import time

def extract_data_from_txt(txt_file):
    data = []
    with open(txt_file, "r", encoding="utf-8") as file:
        for line in file:
            code, account = line.split(maxsplit=1)
            code = code.ljust(9, '0')
            code = '.'.join(code[i:i+3] for i in range(0, len(code), 3))
            data.append((code, account)) 
            print(code, account)  
            time.sleep(0.2)  
    return data

def write_to_excel(data, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row, (code, account) in enumerate(data, start=1):
        sheet.cell(row=row, column=1).value = code
        sheet.cell(row=row, column=2).value = account
    
    workbook.save(excel_file)

if __name__ == "__main__":
    txt_file = "1st.txt"
    excel_file = "planDeCuentas.xlsx"
    data = extract_data_from_txt(txt_file)
    write_to_excel(data, excel_file)